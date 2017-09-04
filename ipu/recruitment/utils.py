from student.models import Student
import openpyxl as excel, time
from datetime import date
from hashids import Hashids

def get_excel_structure(main_heading, secondary_heading, students_queryset):
	workbook = excel.Workbook()
	worksheet = workbook.active
	worksheet.title = "Placement Session"
	worksheet['A1'].font = excel.styles.Font(name='Times New Roman', size=20, bold=True)
	worksheet.merge_cells("A1:I2");worksheet.merge_cells("A3:I3")
	worksheet['A1'] = main_heading # Job/Internship @ College
	worksheet['A3'] = secondary_heading # Programme - Streams
	worksheet.freeze_panes = 'A5'
	
	# S.No. | Enrollment No. | First Name | Last Name | Gender | Email | Stream | Year | Tenth | Twelfth | Graduation | Post Grad | Doctorate
	worksheet['A4'] = 'S.No.'; worksheet['B4'] = 'Enrollment No.'; worksheet['C4'] = 'First Name'; worksheet['D4'] = 'Last Name';
	worksheet['E4'] = 'Gender'; worksheet['F4'] = 'Email'; worksheet['G4'] = 'Stream'; worksheet['H4'] = 'Year';
	worksheet['I4'] = '10th'; worksheet['J4'] = '12th'; worksheet['K4'] = 'Graduation'; worksheet['L4'] = 'Post Graduation'; worksheet['M4'] = 'Doctorate'
	
	bold = excel.styles.Font(bold=True)
	for i in range(1,14):
		worksheet.cell(row=4, column=i).font = bold
	
	GENDER = dict(Student.GENDER_CHOICES)
	for i, student in enumerate(students_queryset, 1):
		qualifications = getattr(student, 'qualifications', None)
		row = worksheet.max_row+1
		worksheet.cell(row=row, column=1).value = i
		worksheet.cell(row=row, column=2).value = student.profile.username
		worksheet.cell(row=row, column=3).value = student.firstname.title()
		worksheet.cell(row=row, column=4).value = student.lastname.title()
		worksheet.cell(row=row, column=5).value = GENDER[student.gender].__str__()
		worksheet.cell(row=row, column=6).value = student.profile.email
		worksheet.cell(row=row, column=7).value = student.stream.name.title()
		worksheet.cell(row=row, column=8).value = student.current_year
		worksheet.cell(row=row, column=9).value = get_qual_value(qualifications, 'tenth')
		worksheet.cell(row=row, column=10).value = get_qual_value(qualifications, 'twelfth')
		worksheet.cell(row=row, column=11).value = get_qual_value(qualifications, 'graduation')
		worksheet.cell(row=row, column=12).value = get_qual_value(qualifications, 'post_graduation')
		worksheet.cell(row=row, column=13).value = get_qual_value(qualifications, 'doctorate')
	
	to_letter = excel.cell.get_column_letter
	for col in [1,5,8]:
		worksheet.column_dimensions[to_letter(col)].width = 5
	for col in [2,3,4,9,10,11,12,13]:
		worksheet.column_dimensions[to_letter(col)].width = 12
	for col in [6,7]:
		worksheet.column_dimensions[to_letter(col)].width = 20
	return workbook

def get_master_excel_structure(college, students_queryset): # All students' data
	students_queryset = students_queryset.order_by('stream__code') # Grouping stream-wise
	workbook = excel.Workbook()
	worksheet = workbook.active
	worksheet.title = "Master"
	worksheet['A1'].font = excel.styles.Font(name='Times New Roman', size=20, bold=True)
	worksheet.merge_cells("A1:I2");worksheet.merge_cells("A3:I3")
	worksheet['A1'] = college.name.title() # College Name
	worksheet['A3'] = date.today().strftime("Students' data as of %d %b, %Y") # Date
	worksheet.freeze_panes = 'A5'
	
	# S.No. | Enrollment No. | First Name | Last Name | Gender | Email | Programme | Stream | Year
	# | Tenth CGPA | Tenth CC | Tenth Board | Tenth %
	# | Twelfth Board | Twelfth %
	# | Graduation | Post Grad | Doctorate
	# | Verified | Back(s) | Min. Salary Expected
	# | Phone Number | DoB         -- Not serving these because of privacy stipulation

	worksheet['A4'] = 'S.No.'; worksheet['B4'] = 'Enrollment No.'; worksheet['C4'] = 'First Name'; worksheet['D4'] = 'Last Name';
	worksheet['E4'] = 'Gender'; worksheet['F4'] = 'Email'; worksheet['G4'] = 'Programme'; worksheet['H4'] = 'Stream'; worksheet['I4'] = 'Year';
	worksheet['J4'] = '10th CGPA'; worksheet['K4'] = '10th Conversion Factor'; worksheet['L4'] = '10th Board'; worksheet['M4'] = '10th %';
	worksheet['N4'] = '12th Board'; worksheet['O4'] = '12th %'; worksheet['P4'] = 'Graduation'; worksheet['Q4'] = 'Post Graduation';
	worksheet['R4'] = 'Doctorate'; worksheet['S4'] = 'Verified'; worksheet['T4'] = 'Back(s)'; worksheet['U4'] = 'Min. Salary Expected';
#	worksheet['V4'] = 'Phone Number'; worksheet['W4'] = 'Date of Birth';
	
	bold = excel.styles.Font(bold=True)
	for i in range(1,24):
		worksheet.cell(row=4, column=i).font = bold
	
	GENDER = dict(Student.GENDER_CHOICES)
	for i, student in enumerate(students_queryset, 1):
		qualifications = getattr(student, 'qualifications', None)
		row = worksheet.max_row+1
		worksheet.cell(row=row, column=1).value = i
		worksheet.cell(row=row, column=2).value = student.profile.username
		worksheet.cell(row=row, column=3).value = student.firstname.title()
		worksheet.cell(row=row, column=4).value = student.lastname.title()
		worksheet.cell(row=row, column=5).value = GENDER[student.gender].__str__()
		worksheet.cell(row=row, column=6).value = student.profile.email

		worksheet.cell(row=row, column=7).value = student.programme.name
		worksheet.cell(row=row, column=8).value = student.stream.name.title()
		worksheet.cell(row=row, column=9).value = student.current_year
		
		worksheet.cell(row=row, column=10).value, worksheet.cell(row=row, column=11).value = get_tenth_cgpa(student)
		worksheet.cell(row=row, column=12).value = get_board(student, '10')
		worksheet.cell(row=row, column=13).value = get_qual_value(qualifications, 'tenth')

		worksheet.cell(row=row, column=14).value = get_board(student, '12')
		worksheet.cell(row=row, column=15).value = get_qual_value(qualifications, 'twelfth')

		worksheet.cell(row=row, column=16).value = get_qual_value(qualifications, 'graduation')
		worksheet.cell(row=row, column=17).value = get_qual_value(qualifications, 'post_graduation')
		worksheet.cell(row=row, column=18).value = get_qual_value(qualifications, 'doctorate')

		worksheet.cell(row=row, column=19).value = "Yes" if (student.is_verified and student.verified_by) else "No"
		worksheet.cell(row=row, column=20).value = "Yes" if student.is_sub_back else "No"
		worksheet.cell(row=row, column=21).value = ("%d LPA" % student.salary_expected) if student.salary_expected else "--"
#		worksheet.cell(row=row, column=22).value = student.phone_number
#		worksheet.cell(row=row, column=23).value = student.dob.__str__()
	
	to_letter = excel.cell.get_column_letter
	for col in [1,5,9,19,20,23]:
		worksheet.column_dimensions[to_letter(col)].width = 7
	for col in [2,3,4,7,10,11,12,13,14,15,16,17,18,22]:
		worksheet.column_dimensions[to_letter(col)].width = 12
	for col in [6,8,21]:
		worksheet.column_dimensions[to_letter(col)].width = 20
	return workbook


def get_qual_value(qualifications, attr):
	value = getattr(qualifications, attr, None)
	if not value or float(value) == 33:
		return '--'
	else:
		return str(value) + '%'

def get_tenth_cgpa(student):
	try:
		marksheet = student.marksheet
		cgpa = marksheet.cgpa_marksheet
		if cgpa:
			return (cgpa.cgpa, cgpa.conversion_factor)
		return ('--','--')
	except:
		return ('--','--')

def get_board(student, klass):
	try:
		marksheet = student.marksheet
		if klass == '10':
			if marksheet.cgpa_marksheet:
				board = marksheet.cgpa_marksheet.board
				return board.abbreviation if board.abbreviation else board.name
			else:
				board = marksheet.marksheet_10.board
				return board.abbreviation if board.abbreviation else board.name
		else:
			board = marksheet.marksheet_12.board
			return board.abbreviation if board.abbreviation else board.name
	except:
		return '--'
